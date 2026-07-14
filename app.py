import streamlit as st
import pandas as pd
import io
import base64
import os
from PIL import Image
from openpyxl.styles import Font, numbers, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# --- Handle Logo Loading ---
logo_path = "/content/download.jfif"
try:
    with open(logo_path, "rb") as f:
        logo_base64 = base64.b64encode(f.read()).decode('utf-8')
    logo_html = f'<img src="data:image/jpeg;base64,{logo_base64}" style="height: 40px; vertical-align: middle; margin-right: 15px; border-radius: 5px;">'
except Exception:
    logo_html = "🚗 "

# --- Page Configuration ---
st.set_page_config(
    page_title="Slice Parking Payouts",
    page_icon="🅿️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- Custom Styling ---
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');

    html, body, [class*='css'] {
        font-family: 'Inter', sans-serif;
    }
    .main-header {font-size: 2.4rem; font-weight: 700; color: #1E3A8A; margin-bottom: 5px; display: flex; align-items: center;}
    .sub-header {font-size: 1.1rem; font-weight: 500; color: #475569; margin-bottom: 25px;}
    .stButton>button {background-color: #1E3A8A; color: white; border-radius: 8px; font-weight: 600; height: 50px; transition: all 0.3s ease; border: none; box-shadow: 0 4px 6px -1px rgba(30, 58, 138, 0.2);}
    .stButton>button:hover {background-color: #1E40AF; color: white; box-shadow: 0 10px 15px -3px rgba(30, 58, 138, 0.3); transform: translateY(-2px);}

    /* Modern Metric Cards */
    .metric-card {background: #FFFFFF; padding: 24px; border-radius: 12px; border: 1px solid #E2E8F0; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05); text-align: center;}
    .metric-title {font-size: 0.95rem; font-weight: 600; color: #475569; text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 8px;}
    .metric-value {font-size: 2rem; font-weight: 700; color: #065F46; margin: 0;}
    .metric-highlight {color: #047857;}

    .section-title {font-size: 1.5rem; font-weight: 700; color: #1E3A8A; margin-top: 30px; margin-bottom: 15px; border-bottom: 2px solid #E2E8F0; padding-bottom: 5px;}
    </style>
""", unsafe_allow_html=True)

# --- Main Dashboard ---
st.markdown(f"<div class='main-header'>{logo_html} Slice Parking Payout Generator</div>", unsafe_allow_html=True)
st.markdown("<div class='sub-header'>Professional financial reconciliation & automated payout reporting.</div>", unsafe_allow_html=True)

st.markdown("<div class='section-title'>⚙️ Step 1: Configuration</div>", unsafe_allow_html=True)

# Initialize session state for sliders if not present
if 'slice_pct' not in st.session_state:
    st.session_state.slice_pct = 30
if 'client_pct' not in st.session_state:
    st.session_state.client_pct = 70

def update_client():
    st.session_state.client_pct = 100 - st.session_state.slice_pct

def update_slice():
    st.session_state.slice_pct = 100 - st.session_state.client_pct

col0, col1, col2, col3 = st.columns(4)
with col0:
    parking_code = st.selectbox("🅿️ Parking Code", ["BCW", "CP", "TAJ", "GH", "Gator", "HBH", "CHS", "BDCHP", "Billiards", "BlueWave", "Crain", "Dyson", "EastonDemo", "SunBeach", "UUMC"])
with col1:
    target_date = st.date_input("📅 Target Payout Date", value=pd.to_datetime('2026-07-03').date())
with col2:
    slice_pct = st.slider("🍕 Slice %", min_value=0, max_value=100, key="slice_pct", on_change=update_client)
with col3:
    client_pct = st.slider("🤝 Client %", min_value=0, max_value=100, key="client_pct", on_change=update_slice)

if slice_pct + client_pct != 100:
    st.warning("⚠️ Heads up: Your Slice and Client percentages do not add up to 100%.")

# --- Sidebar Inputs ---
with st.sidebar:
    st.markdown(f"<div style='text-align: center; margin-bottom: 20px; font-size: 50px;'>🅿️</div>", unsafe_allow_html=True)
    st.markdown("## 📂 Upload Data Files")

    if parking_code == "HBH":
        main_file = st.file_uploader("1. Main Transaction Data (HBH)", type=["xlsx", "xls"], key="hbh_main")
        enf_file = st.file_uploader("2. Enforcement Data (Payout)", type=["xlsx", "xls"], key="hbh_enf")
        tier_file = st.file_uploader("3. Tier Data (Parkpliant)", type=["xlsx", "xls"], key="hbh_tier")
        sh_file = st.file_uploader("4. Spot Hero Data (HBH Only)", type=["csv"], key="hbh_sh")
    elif parking_code in ["CHS", "BDCHP", "Billiards", "BlueWave", "Crain", "Dyson", "EastonDemo", "SunBeach", "UUMC"]:
        main_file = st.file_uploader("1. Main Transaction Data (Slice Parking)", type=["xlsx", "xls"], key="std_main")
        sh_file = None
        enf_file = None
        tier_file = None
    elif parking_code in ["BCW", "CP", "TAJ"]:
        main_file = st.file_uploader("1. Main Transaction Data (Slice Parking)", type=["xlsx", "xls"], key="std_main")
        enf_file = st.file_uploader("2. Enforcement Data (Payout)", type=["xlsx", "xls"], key="std_enf")
        tier_file = st.file_uploader("3. Tier Data (Parkpliant)", type=["xlsx", "xls"], key="std_tier")
        sh_file = None
    else: # GH, Gator
        main_file = st.file_uploader("1. Main Transaction Data (Slice Parking)", type=["xlsx", "xls"], key="std_main")
        enf_file = st.file_uploader("2. Enforcement Data (Payout)", type=["xlsx", "xls"], key="std_enf")
        tier_file = st.file_uploader("3. Tier Data (Parkpliant)", type=["xlsx", "xls"], key="std_tier")
        sh_file = st.file_uploader("4. Spot Hero Data (CSV)", type=["csv"], key="std_sh")

    st.markdown("---")
    st.markdown("**Need help?** Contact your Slice Parking admin.")

st.markdown("<div class='section-title'>🚀 Step 2: Generate Report</div>", unsafe_allow_html=True)
generate_btn = st.button("Generate Payout Report", use_container_width=True)

# --- Main Processing Logic ---
if generate_btn:
    if parking_code == "HBH":
        required_files = [main_file, enf_file, tier_file, sh_file]
    elif parking_code in ["CHS", "BDCHP", "Billiards", "BlueWave", "Crain", "Dyson", "EastonDemo", "SunBeach", "UUMC"]:
        required_files = [main_file]
    else:
        required_files = [main_file, enf_file, tier_file]
        if parking_code in ["GH", "Gator"]:
            required_files.append(sh_file)

    if not all(required_files):
        st.error(f"🛑 Please upload all required files for {parking_code} in the sidebar.")
    else:
        with st.spinner("Crunching the numbers... Please wait..."):
            try:
                # 1. Load Data
                df = pd.read_excel(main_file)
                er = pd.read_excel(enf_file) if enf_file else pd.DataFrame()
                pp = pd.read_excel(tier_file) if tier_file else pd.DataFrame()
                sh = pd.read_csv(sh_file) if sh_file else pd.DataFrame()

                # 2. Logic configuration per code
                target_date_pd = pd.to_datetime(target_date).date()
                if 'Payout Date' in df.columns:
                    df['Payout Date'] = pd.to_datetime(df['Payout Date'], errors='coerce').dt.date

                Enforce = 0.0
                sh_sum = 0.0
                net_calc = 0.0
                total_net = 0.0
                payout = 0.0
                specific_sum_pp = 0.0

                if not pp.empty:
                    T1, T2, T3, Te, L1, L2 = 1.5, 2, 2, 1, 2, 2
                    pp_sum_series = (pp['Tier 1 Lookup']*T1 + pp['Tier 2 Lookup']*T2 +
                                     pp['Tier 3 Lookup']*T3 + pp['Text 1']*Te +
                                     pp['Letter 1']*L1 + pp['Letter 2']*L2)
                else:
                    pp_sum_series = pd.Series()

                if parking_code == "HBH":
                    df_main = df[(pd.to_datetime(df['Payout Date'], errors='coerce').dt.date == target_date_pd)]
                else:
                    df_main = df[(df['Parking Code'] == parking_code) & (pd.to_datetime(df['Payout Date'], errors='coerce').dt.date == target_date_pd)]

                gross = float(df_main['Total Amount'].sum()) if not df_main.empty else 0.0
                service = float(df_main['Service Fee'].sum()) if not df_main.empty else 0.0
                credit = float(df_main['Payment Gateway Fee'].sum()) if 'Payment Gateway Fee' in df_main.columns else 0.0


                if parking_code == "BCW":
                  tax = float(df_main['City Tax'].sum()) if not df_main.empty else 0.0
                  subtotal = float(gross - service - credit - tax)
                else:
                  subtotal = float(gross - service - credit)


                if parking_code in ["HBH", "Billiards", "Dyson"]:
                    dock_pct = 0.05
                    dock_pct_str = "5%"
                else:
                    dock_pct = 0.10
                    dock_pct_str = "10%"

                dock = float(subtotal * dock_pct)
                net = float(subtotal - dock)

                if parking_code == "BCW":
                    filtered_er = er[er['Lot Code'] == 'Baltimore Clayworks']
                    sum_net_enf = float(filtered_er['Net'].sum()) if not filtered_er.empty else 0.0
                    specific_sum_pp = float(pp_sum_series.iloc[1]) if len(pp_sum_series) > 1 else 0.0
                    Enforce = float(sum_net_enf - specific_sum_pp)

                    bcws_all_rows = df[
                        (df['Parking Code'].astype(str).str.strip().str.upper() == 'BCWS') &
                        (df['Payout Date'] == target_date_pd) &
                        (df['Purpose'].astype(str).str.strip().str.upper() == 'SUBSCRIPTION')
                    ]
                    if not bcws_all_rows.empty:
                        total_calc = float(bcws_all_rows['Total Amount'].sum() - bcws_all_rows['Payment Gateway Fee'].sum() - bcws_all_rows['Service Fee'].sum() - bcws_all_rows['City Tax'].sum())
                        net_calc = float(total_calc - (total_calc * 0.10))
                    else:
                        total_calc = 0.0
                        net_calc = 0.0

                    total_net = float(net + net_calc + Enforce)
                    payout = float(total_net * (client_pct / 100))

                elif parking_code == "CP":
                    filtered_er = er[er['Lot Code'] == 'CP']
                    sum_net_enf = float(filtered_er['Net'].sum()) if not filtered_er.empty else 0.0
                    specific_sum_pp = float(pp_sum_series.iloc[3]) if len(pp_sum_series) > 3 else 0.0 # Assuming index 3 for CP, you might need to adjust based on exact Parkpliant data
                    Enforce = float(sum_net_enf - specific_sum_pp)

                    cp_upper_rows = df[
                        (df['Parking Code'].astype(str).str.strip() == 'CPUpperLot') &
                        (df['Payout Date'] == target_date_pd) &
                        (df['Purpose'].astype(str).str.strip().str.upper() == 'PARKING')
                    ]
                    if not cp_upper_rows.empty:
                        total_calc = float(cp_upper_rows['Total Amount'].sum() - cp_upper_rows['Payment Gateway Fee'].sum() - cp_upper_rows['Service Fee'].sum())
                        net_calc = float(total_calc - (total_calc * 0.10))
                    else:
                        total_calc = 0.0
                        net_calc = 0.0

                    total_net = float(net + net_calc + Enforce)
                    payout = float(total_net * (client_pct / 100))

                elif parking_code == "TAJ":
                    filtered_er = er[er['Lot Code'] == 'TAJ']
                    sum_net_enf = float(filtered_er['Net'].sum()) if not filtered_er.empty else 0.0
                    specific_sum_pp = float(pp_sum_series.iloc[4]) if len(pp_sum_series) > 4 else 0.0 # Assuming index 4 for TAJ
                    Enforce = float(sum_net_enf - specific_sum_pp)

                    total_net = float(net + Enforce)
                    payout = float(total_net * (client_pct / 100))

                elif parking_code == "GH":
                    filtered_er = er[er['Lot Code'] == 'Great Harvest Annapolis Lot']
                    sum_net_enf = float(filtered_er['Net'].sum()) if not filtered_er.empty else 0.0
                    specific_sum_pp = float(pp_sum_series.iloc[0]) if len(pp_sum_series) > 0 else 0.0
                    Enforce = float(sum_net_enf - specific_sum_pp)

                    if not sh.empty:
                        filtered_sh = sh[sh['spot'] == '208 Ridgely Ave. - Great Harvest Annapolis Lot']
                        sh_sum = float(filtered_sh['total remit'].sum()) if not filtered_sh.empty else 0.0

                    total_net = float(net + sh_sum + Enforce)
                    payout = float(total_net * (client_pct / 100))

                elif parking_code == "Gator":
                    if not sh.empty:
                        filtered_sh = sh[sh['spot'] == '999 Anastasia Blvd. - Alligator Farm Lot']
                        sh_sum = float(filtered_sh['total remit'].sum()) if not filtered_sh.empty else 0.0

                    total_net = float(net + sh_sum)
                    payout = float(total_net * (client_pct / 100))

                elif parking_code == "HBH":
                    hbh_lots = [
                        'Cleveland Lot HBH', 'Del Mar Lot', 'McKinley Street Lot',
                        'Nebraska Lot - HBH', 'Neptune Lot', 'Paradise Lot',
                        'Seaside Lot', 'SILVER SPRAY  Lot', 'The Swan Lot'
                    ]
                    filtered_er = er[er['Lot Code'].isin(hbh_lots)]
                    sum_net_enf = float(filtered_er['Net'].sum()) if not filtered_er.empty else 0.0
                    specific_sum_pp = float(pp_sum_series.iloc[2]) if len(pp_sum_series) > 2 else 0.0
                    Enforce = float(sum_net_enf - specific_sum_pp)

                    if not sh.empty:
                        if 'total remit' in sh.columns:
                            sh_sum = float(sh['total remit'].sum())
                        else:
                            remit_col = [col for col in sh.columns if 'total remit' in col.lower()]
                            if remit_col:
                                sh_sum = float(sh[remit_col[0]].sum())

                    total_net = float(net + sh_sum + Enforce)
                    payout = float(total_net * (client_pct / 100))

                elif parking_code in ["CHS", "BDCHP", "Billiards", "BlueWave", "Crain", "Dyson", "EastonDemo", "SunBeach", "UUMC"]:
                    total_net = float(net)
                    payout = float(total_net * (client_pct / 100))

                st.toast("🎉 Success! Report generated.")
                st.success(f"🎉 Success! Financial analysis generated for {parking_code} on {target_date}.")

                def fmt_currency(val):
                    try:
                        v = float(val)
                        if pd.isna(v): return "$0.00"
                        return f"${v:,.2f}"
                    except: return "$0.00"

                st.markdown("<div class='section-title'>📊 Financial Overview</div>", unsafe_allow_html=True)

                # Render Custom Metrics Cards
                st.markdown(f"""
                <div style='display: flex; gap: 20px; flex-wrap: wrap; margin-bottom: 20px;'>
                    <div class='metric-card' style='flex: 1; min-width: 150px;'>
                        <div class='metric-title'>Gross Revenue</div>
                        <div class='metric-value'>{fmt_currency(gross)}</div>
                    </div>
                    <div class='metric-card' style='flex: 1; min-width: 150px;'>
                        <div class='metric-title'>Tech Fee ({dock_pct_str})</div>
                        <div class='metric-value'>{fmt_currency(dock)}</div>
                    </div>
                    <div class='metric-card' style='flex: 1; min-width: 150px;'>
                        <div class='metric-title'>Spot Hero Rev</div>
                        <div class='metric-value'>{fmt_currency(sh_sum)}</div>
                    </div>
                    <div class='metric-card' style='flex: 1; min-width: 150px;'>
                        <div class='metric-title'>Taggr Rev</div>
                        <div class='metric-value'>{fmt_currency(Enforce)}</div>
                    </div>
                    <div class='metric-card' style='flex: 1; min-width: 150px; border-bottom: 4px solid #065F46;'>
                        <div class='metric-title'>Payout ({client_pct}% Split)</div>
                        <div class='metric-value metric-highlight'>{fmt_currency(payout)}</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                st.markdown("<h5 style='font-weight: 700; color: #1E3A8A; margin-top: 20px;'>Detailed Summary Breakdown</h5>", unsafe_allow_html=True)
                if parking_code == "BCW":
                  metrics = ['Gross Revenue', 'Service Fee', 'Payment Gateway Fee', 'City Tax', 'Subtotal', 'Tech Fee', 'Parking Net Revenue']
                  values = [gross, service, credit, tax, subtotal, dock, net]
                else:
                  metrics = ['Gross Revenue', 'Service Fee', 'Payment Gateway Fee', 'Subtotal', 'Tech Fee', 'Parking Net Revenue']
                  values = [gross, service, credit, subtotal, dock, net]

                if parking_code == 'BCW':
                    metrics.extend(['Parkpliant Revenue', 'Taggr Revenue', 'Subscription Net Revenue'])
                    values.extend([specific_sum_pp, Enforce, net_calc])
                elif parking_code == 'CP':
                    metrics.extend(['Parkpliant Revenue', 'Taggr Revenue', 'Subscription Net Revenue'])
                    values.extend([specific_sum_pp, Enforce, net_calc])
                elif parking_code == 'TAJ':
                    metrics.extend(['Parkpliant Revenue', 'Taggr Revenue'])
                    values.extend([specific_sum_pp, Enforce])
                elif parking_code == 'GH':
                    metrics.extend(['Parkpliant Revenue', 'Taggr Revenue', 'Spot Hero Revenue'])
                    values.extend([specific_sum_pp, Enforce, sh_sum])
                elif parking_code == 'HBH':
                    metrics.extend(['Parkpliant Revenue', 'Taggr Revenue', 'Spot Hero Revenue'])
                    values.extend([specific_sum_pp, Enforce, sh_sum])
                elif parking_code in ['Gator']:
                    metrics.extend(['Spot Hero Revenue'])
                    values.extend([sh_sum])

                metrics.extend(['Total Net Revenue', f'Total Payout ({client_pct}%)'])
                values.extend([total_net, payout])

                main_summary_ui = pd.DataFrame({'Metric': metrics, 'Value': values})
                st.dataframe(main_summary_ui.style.format({"Value": "${:,.2f}"}), use_container_width=True)

                if parking_code == 'BCW':
                  columns_to_drop = ['Application Fee', 'Month', 'Time', 'Date', 'County Tax', 'Owner Payout', 'Client Payout', 'Operator Payout', 'Dock Revenue', 'Purpose', 'id', 'Total Refund', 'Refund Transaction ID', 'Customer ID', 'Payment Method ID', 'Transaction ID', 'Year', 'Weekday', 'Hour', 'Payout Date', 'Payout Difference', 'Difference', 'Failed Reason', 'Start Date', 'End Date', 'Rate', 'Extend Reminder Sent?', 'Extended Reservation', 'Validation Code', 'Email', 'Name', 'Space Number']
                else:
                  columns_to_drop = ['Application Fee', 'Month', 'Time', 'Date', 'County Tax', 'Owner Payout', 'Client Payout', 'Operator Payout', 'Dock Revenue', 'Purpose', 'id', 'Total Refund', 'Refund Transaction ID', 'Customer ID', 'Payment Method ID', 'Transaction ID', 'Year', 'Weekday', 'Hour', 'Payout Date', 'Payout Difference', 'Difference', 'Failed Reason', 'Start Date', 'End Date', 'Rate', 'Extend Reminder Sent?', 'Extended Reservation', 'Validation Code', 'Email', 'Name', 'Space Number', 'City Tax']
                cols_to_drop = [c for c in columns_to_drop if c in df_main.columns]
                df_export = df_main.drop(columns=cols_to_drop).copy()

                if 'Transaction Date' in df_export.columns:
                    parsed_dates = pd.to_datetime(df_export['Transaction Date'], errors='coerce')
                    df_export['Transaction Date'] = parsed_dates.dt.strftime('%Y-%m-%d %H:%M:%S').fillna(df_export['Transaction Date'])

                # Create summary with a blank column in between
                main_summary_export = pd.DataFrame({'': metrics, ' ': [''] * len(metrics), '  ': values})

                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df_export.to_excel(writer, sheet_name='Report', index=False, startrow=0)

                    workbook = writer.book
                    worksheet = writer.sheets['Report']

                    # Freeze the first header row
                    worksheet.freeze_panes = 'A2'

                    # --- Format Main Data as Excel Table ---
                    end_col_letter = get_column_letter(len(df_export.columns))
                    table_ref = f"A1:{end_col_letter}{len(df_export) + 1}"
                    tab = Table(displayName="MainTable", ref=table_ref)

                    # Add a default style with striped rows and border lines
                    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                    tab.tableStyleInfo = style
                    worksheet.add_table(tab)

                    # Identify columns to format as currency in main table
                    currency_cols = ['Base Rate', 'Tax', 'Service Fee', 'City Tax', 'Payment Gateway Fee', 'Total Amount']
                    col_indices = {col_name: idx for idx, col_name in enumerate(df_export.columns, 1)}

                    # Apply currency format to main table
                    for col_name in currency_cols:
                        if col_name in col_indices:
                            col_idx = col_indices[col_name]
                            for row in range(2, len(df_export) + 2):
                                worksheet.cell(row=row, column=col_idx).number_format = '"$"#,##0.00'

                    # --- Write Main Summary ---
                    # GAP of 6 rows added
                    summary_start_row = len(df_export) + 2 + 6
                    main_summary_export.to_excel(writer, sheet_name='Report', index=False, header=False, startrow=summary_start_row - 1)

                    # Format summary section
                    bold_metrics = ['Gross Revenue', 'Subtotal', 'Parking Net Revenue', 'Total Net Revenue']
                    for row in range(summary_start_row, summary_start_row + len(main_summary_export)):
                        metric_cell = worksheet.cell(row=row, column=1)
                        value_cell = worksheet.cell(row=row, column=3)

                        # Currency format for calculation values
                        value_cell.number_format = '"$"#,##0.00'

                        # Bold specific metrics
                        if str(metric_cell.value) in bold_metrics or str(metric_cell.value).startswith('Total Payout'):
                            metric_cell.font = Font(bold=True)
                            value_cell.font = Font(bold=True)

                        # Green color for Total Payout number
                        if str(metric_cell.value).startswith('Total Payout'):
                            value_cell.font = Font(bold=True, color="008000")

                    # Make it more attractive: adjust column widths automatically
                    for col in worksheet.columns:
                        max_length = 0
                        column_letter = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[column_letter].width = adjusted_width

                    if parking_code == 'BCW' and 'bcws_all_rows' in locals() and not bcws_all_rows.empty:
                        bcws_export = bcws_all_rows.drop(columns=[c for c in columns_to_drop if c in bcws_all_rows.columns]).copy()
                        sub_summary_data = {'': ['Total Amount', 'Payment Gateway Fee', 'Service Fee', 'City Tax', 'Sub Total', 'Tech Fee', 'Net Total'], ' ': [''] * 7, '  ': [float(bcws_all_rows['Total Amount'].sum()), float(bcws_all_rows['Payment Gateway Fee'].sum()), float(bcws_all_rows['Service Fee'].sum()), float(bcws_all_rows['City Tax'].sum()), total_calc, float(total_calc * 0.10), net_calc]}
                        sub_summary = pd.DataFrame(sub_summary_data)

                        bcws_start = summary_start_row + len(main_summary_export) + 2
                        bcws_export.to_excel(writer, sheet_name='Report', index=False, startrow=bcws_start)

                        # --- Format BCWS Data as Excel Table ---
                        bcws_end_col_letter = get_column_letter(len(bcws_export.columns))
                        bcws_table_ref = f"A{bcws_start + 1}:{bcws_end_col_letter}{bcws_start + len(bcws_export) + 1}"
                        tab_bcws = Table(displayName="BCWSTable", ref=bcws_table_ref)
                        tab_bcws.tableStyleInfo = style
                        worksheet.add_table(tab_bcws)

                        # 6 row gap before BCWS summary
                        bcws_summary_row_idx = bcws_start + len(bcws_export) + 1 + 6
                        sub_summary.to_excel(writer, sheet_name='Report', index=False, header=False, startrow=bcws_summary_row_idx - 1)

                        for row in range(bcws_summary_row_idx, bcws_summary_row_idx + len(sub_summary)):
                             worksheet.cell(row=row, column=3).number_format = '"$"#,##0.00'
                    
                    elif parking_code == 'CP' and 'cp_upper_rows' in locals() and not cp_upper_rows.empty:
                        cp_export = cp_upper_rows.drop(columns=[c for c in columns_to_drop if c in cp_upper_rows.columns]).copy()
                        sub_summary_data = {'': ['Total Amount', 'Payment Gateway Fee', 'Service Fee', 'Sub Total', 'Tech Fee', 'Net Total'], ' ': [''] * 5, '  ': [float(cp_upper_rows['Total Amount'].sum()), float(cp_upper_rows['Payment Gateway Fee'].sum()), float(cp_upper_rows['Service Fee'].sum()), total_calc, float(total_calc * 0.10), net_calc]}
                        sub_summary = pd.DataFrame(sub_summary_data)

                        cp_start = summary_start_row + len(main_summary_export) + 2
                        cp_export.to_excel(writer, sheet_name='Report', index=False, startrow=cp_start)

                        # --- Format CPUpper Data as Excel Table ---
                        cp_end_col_letter = get_column_letter(len(cp_export.columns))
                        cp_table_ref = f"A{cp_start + 1}:{cp_end_col_letter}{cp_start + len(cp_export) + 1}"
                        tab_cp = Table(displayName="CPUpperTable", ref=cp_table_ref)
                        tab_cp.tableStyleInfo = style
                        worksheet.add_table(tab_cp)

                        # 6 row gap before CPUpper summary
                        cp_summary_row_idx = cp_start + len(cp_export) + 1 + 6
                        sub_summary.to_excel(writer, sheet_name='Report', index=False, header=False, startrow=cp_summary_row_idx - 1)

                        for row in range(cp_summary_row_idx, cp_summary_row_idx + len(sub_summary)):
                             worksheet.cell(row=row, column=3).number_format = '"$"#,##0.00'

                st.markdown("---")
                st.markdown("### 📥 Export")
                st.download_button(label=f"⬇️ Download {parking_code} Report (Excel)", data=buffer.getvalue(), file_name=f"{parking_code}_Payout_Report_{target_date}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")

            except Exception as e:
                st.error(f"An error occurred during processing: {e}")
